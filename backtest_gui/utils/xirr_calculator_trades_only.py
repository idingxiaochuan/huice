#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
交易专用XIRR计算模块 - 只计算实际交易的年化收益率，不考虑初始资金
"""
import numpy as np
import pandas as pd
from datetime import datetime
from scipy import optimize
import traceback

class XIRRCalculatorTradesOnly:
    """交易专用XIRR计算器"""
    
    def __init__(self, db_connector):
        """初始化XIRR计算器
        
        Args:
            db_connector: 数据库连接器
        """
        self.db_connector = db_connector
    
    def _xnpv(self, rate, values, dates):
        """计算XNPV（净现值）- 与Excel完全一致的实现
        
        Args:
            rate: 折现率
            values: 现金流金额列表
            dates: 现金流日期列表
            
        Returns:
            XNPV值
        """
        if rate <= -1.0:
            return float('inf')
        
        # 确保日期是datetime对象
        dates = [pd.to_datetime(dt) if not isinstance(dt, datetime) else dt for dt in dates]
        
        # 确保第一个日期是最早的日期
        min_date = min(dates)
        
        # 计算每个现金流的折现值 - 使用与Excel完全相同的算法
        result = 0.0
        for i in range(len(values)):
            days = (dates[i] - min_date).days
            result += values[i] / pow(1.0 + rate, days / 365.0)
        
        return result
    
    def _xirr_objective(self, rate, values, dates):
        """XIRR优化目标函数
        
        Args:
            rate: 折现率
            values: 现金流金额列表
            dates: 现金流日期列表
            
        Returns:
            XNPV值
        """
        return self._xnpv(rate, values, dates)
    
    def calculate_xirr(self, dates, amounts, guess=0.06):
        """计算XIRR - 与Excel完全一致的实现
        
        Args:
            dates: 日期列表
            amounts: 金额列表
            guess: 初始猜测值，默认使用0.06接近Excel结果
            
        Returns:
            XIRR值
        """
        try:
            # 检查输入
            if len(dates) != len(amounts):
                raise ValueError("日期和金额列表长度不一致")
            
            if len(dates) < 2:
                raise ValueError("至少需要两个现金流才能计算XIRR")
            
            # 检查现金流是否有效
            if sum(amounts) == 0:
                raise ValueError("现金流总和为零，无法计算XIRR")
            
            # 检查是否有正负现金流
            if all(amount > 0 for amount in amounts) or all(amount < 0 for amount in amounts):
                raise ValueError("XIRR计算需要同时有正负现金流")
            
            # 确保日期和金额按日期排序
            sorted_indices = sorted(range(len(dates)), key=lambda i: dates[i])
            dates = [dates[i] for i in sorted_indices]
            amounts = [amounts[i] for i in sorted_indices]
            
            # 使用Excel默认的初始猜测值0.1
            try:
                result = optimize.newton(
                    lambda r: self._xnpv(r, amounts, dates),
                    guess,
                    tol=0.0000001,  # 提高精度以匹配Excel
                    maxiter=100
                )
                if -1.0 < result < 100.0:  # 合理的XIRR范围
                    return result
            except:
                # 如果Newton方法失败，尝试不同的初始猜测值
                guesses = [0.01, 0.05, 0.1, 0.2, -0.1, -0.05]
                for g in guesses:
                    try:
                        result = optimize.newton(
                            lambda r: self._xnpv(r, amounts, dates),
                            g,
                            tol=0.0000001,
                            maxiter=100
                        )
                        if -1.0 < result < 100.0:
                            return result
                    except:
                        continue
            
            # 如果newton方法失败，尝试brentq方法（Excel在Newton法失败时也会使用类似方法）
            try:
                result = optimize.brentq(
                    lambda r: self._xnpv(r, amounts, dates),
                    -0.999999,  # 避免除以零错误
                    10.0,       # 更合理的上限
                    maxiter=100,
                    xtol=0.0000001  # 提高精度以匹配Excel
                )
                return result
            except:
                pass
            
            # 如果所有方法都失败，返回None
            return None
        except Exception as e:
            print(f"计算XIRR时出错: {str(e)}")
            traceback.print_exc()
            return None
    
    def calculate_backtest_xirr(self, backtest_id):
        """计算指定回测的XIRR"""
        print("\n======== 交易专用XIRR计算器 - 计算回测ID: {0} ========".format(backtest_id))
        
        try:
            # 获取数据库连接
            print("获取数据库连接...")
            conn = self.db_connector.get_connection()
            cursor = conn.cursor()
            
            # 获取回测基本信息
            cursor.execute("""
                SELECT id, stock_code, start_date, end_date, initial_capital, final_capital, total_profit, total_profit_rate
                FROM backtest_results
                WHERE id = %s
            """, (backtest_id,))
            
            backtest_info = cursor.fetchone()
            if not backtest_info:
                print("未找到回测ID: {0}的信息".format(backtest_id))
                self.db_connector.release_connection(conn)
                return None
            
            # 转换为字典
            backtest_info_dict = {
                'id': backtest_info[0],
                'stock_code': backtest_info[1],
                'start_date': backtest_info[2],
                'end_date': backtest_info[3],
                'initial_capital': float(backtest_info[4]),
                'final_capital': float(backtest_info[5]),
                'total_profit': float(backtest_info[6]),
                'total_profit_rate': float(backtest_info[7])
            }
            
            # 获取配对交易记录
            print("获取配对交易记录...")
            cursor.execute("""
                SELECT id, backtest_id, level, grid_type, 
                       buy_time, buy_price, buy_amount, buy_value, 
                       sell_time, sell_price, sell_amount, sell_value,
                       remaining, band_profit, band_profit_rate, status
                FROM backtest_paired_trades 
                WHERE backtest_id = %s
                ORDER BY buy_time
            """, (backtest_id,))
            
            paired_trades = cursor.fetchall()
            if not paired_trades:
                print("未找到回测ID: {0}的配对交易记录".format(backtest_id))
                self.db_connector.release_connection(conn)
                return None
            
            print("找到 {0} 条配对交易记录".format(len(paired_trades)))
            
            # 处理配对交易
            has_incomplete_trades = False
            remaining_shares = 0
            
            print("\n详细交易记录:")
            print("序号\t买入时间\t\t买入价格\t买入数量\t买入金额\t卖出时间\t\t卖出价格\t卖出数量\t卖出金额\t剩余数量\t盈亏\t收益率")
            print("-" * 120)
            
            total_buy_amount = 0
            total_buy_value = 0
            total_sell_amount = 0
            total_sell_value = 0
            total_remaining = 0
            total_profit = 0
            
            # 创建一个列表来跟踪所有未完全卖出的交易
            incomplete_trades = []
            
            for i, trade in enumerate(paired_trades):
                buy_time = trade[4]
                buy_price = float(trade[5])
                buy_amount = int(trade[6])
                buy_value = float(trade[7])
                sell_time = trade[8]
                sell_price = float(trade[9]) if trade[9] is not None else 0
                sell_amount = int(trade[10]) if trade[10] is not None else 0
                sell_value = float(trade[11]) if trade[11] is not None else 0
                remaining = int(trade[12]) if trade[12] is not None else 0
                band_profit = float(trade[13]) if trade[13] is not None else 0
                band_profit_rate = float(trade[14]) if trade[14] is not None else 0
                
                # 计算盈亏
                if sell_amount > 0:
                    # 只计算卖出部分的成本
                    cost_basis = sell_amount * buy_price
                    profit = sell_value - cost_basis
                    profit_rate = (profit / cost_basis) * 100 if cost_basis > 0 else 0
                else:
                    profit = 0
                    profit_rate = 0
                
                # 打印交易详情
                sell_price_str = f"{sell_price:.4f}" if sell_price > 0 else "0.0000"
                print(f"{i+1}\t{buy_time}\t{buy_price:.4f}\t{buy_amount}\t{buy_value:.2f}\t"
                      f"{sell_time if sell_time else 'N/A'}\t{sell_price_str}\t"
                      f"{sell_amount}\t{sell_value:.2f}\t{remaining}\t{profit:.2f}\t{profit_rate:.2f}%")
                
                # 累计统计
                total_buy_amount += buy_amount
                total_buy_value += buy_value
                total_sell_amount += sell_amount
                total_sell_value += sell_value
                total_remaining += remaining
                total_profit += profit
                
                # 检查是否有未完成交易
                if remaining > 0:
                    has_incomplete_trades = True
                    remaining_shares += remaining
                    # 记录未完全卖出的交易
                    incomplete_trades.append({
                        'buy_time': buy_time,
                        'buy_price': buy_price,
                        'remaining': remaining
                    })
                
                # 检查是否有完全未卖出的交易（sell_time为空）
                if sell_time is None:
                    has_incomplete_trades = True
                    # 记录完全未卖出的交易
                    incomplete_trades.append({
                        'buy_time': buy_time,
                        'buy_price': buy_price,
                        'remaining': buy_amount  # 全部买入量都是剩余的
                    })
            
            # 计算底仓成本和总剩余股数（包括所有未完全卖出和完全未卖出的交易）
            total_remaining_shares = 0
            total_remaining_cost = 0
            
            for trade in incomplete_trades:
                total_remaining_shares += trade['remaining']
                total_remaining_cost += trade['remaining'] * trade['buy_price']
            
            avg_remaining_cost = total_remaining_cost / total_remaining_shares if total_remaining_shares > 0 else 0
            print(f"\n底仓信息: 总股数: {total_remaining_shares}, 总成本: {total_remaining_cost:.2f}, 平均成本: {avg_remaining_cost:.4f}")
            
            print("\n交易汇总:")
            print("总买入数量: {0}, 总买入金额: {1:.2f}".format(total_buy_amount, total_buy_value))
            print("总卖出数量: {0}, 总卖出金额: {1:.2f}".format(total_sell_amount, total_sell_value))
            print("剩余股数: {0}".format(total_remaining_shares))
            print("已实现盈亏: {0:.2f}".format(total_profit))
            
            # 准备XIRR计算的现金流
            print("\n准备XIRR计算的现金流...")
            cash_flows = []  # 存储(日期, 金额)对
            buy_cash_flows = []  # 买入现金流
            sell_cash_flows = []  # 卖出现金流
            
            # 添加买入交易的现金流（负值）
            for trade in paired_trades:
                buy_time = trade[4]
                buy_value = float(trade[7])
                
                # 买入是负现金流
                buy_cash_flows.append((buy_time, buy_value))
                cash_flows.append((buy_time, -buy_value))
            
            # 添加卖出交易的现金流（正值）
            for trade in paired_trades:
                sell_time = trade[8]
                sell_value = float(trade[11]) if trade[11] is not None else 0
                
                # 只添加有卖出的交易
                if sell_time and sell_value > 0:
                    sell_cash_flows.append((sell_time, sell_value))
                    cash_flows.append((sell_time, sell_value))
            
            # 计算剩余持仓的价值
            partially_sold_remaining_shares = 0  # 部分卖出后剩余的股数
            unsold_shares = 0  # 完全未卖出的股数
            
            # 计算两种类型的剩余股数
            for trade in paired_trades:
                buy_amount = int(trade[6])
                sell_amount = int(trade[10]) if trade[10] is not None else 0
                remaining = int(trade[12]) if trade[12] is not None else 0
                
                if sell_amount > 0 and remaining > 0:
                    # 部分卖出的交易
                    partially_sold_remaining_shares += remaining
                elif sell_amount == 0:
                    # 完全未卖出的交易
                    unsold_shares += buy_amount
            
            # 如果有剩余持仓，获取最后价格并添加到现金流
            last_price = None
            remaining_value = 0
            
            if total_remaining_shares > 0:
                print(f"\n有剩余股数 {total_remaining_shares}，尝试获取最后价格...")
                
                # 获取最后价格
                
                # 尝试使用最后一次卖出价格
                print("尝试使用最后一次卖出价格...")
                cursor.execute("""
                    SELECT sell_price FROM backtest_paired_trades 
                    WHERE backtest_id = %s AND sell_price IS NOT NULL 
                    ORDER BY sell_time DESC LIMIT 1
                """, (backtest_id,))
                result = cursor.fetchone()
                if result:
                    last_price = float(result[0])
                    print("使用最后一次卖出价格: {0}".format(last_price))
                
                if last_price:
                    # 计算剩余股票的价值
                    remaining_value = total_remaining_shares * last_price
                    print("剩余股票估值: {0:.2f}".format(remaining_value))
                    
                    # 添加到现金流，使用回测结束日期
                    end_date = backtest_info_dict['end_date']
                    
                    # 分别添加两种类型的剩余股票
                    if partially_sold_remaining_shares > 0:
                        partially_sold_value = partially_sold_remaining_shares * last_price
                        cash_flows.append((end_date, partially_sold_value))
                        print(f"添加买入卖出后剩余底仓现金流: {partially_sold_value:.2f} (日期: {end_date})")
                    
                    if unsold_shares > 0:
                        unsold_value = unsold_shares * last_price
                        cash_flows.append((end_date, unsold_value))
                        print(f"添加买入未卖出底仓现金流: {unsold_value:.2f} (日期: {end_date})")
                    
                    # 计算未实现盈亏
                    unrealized_profit = remaining_value - total_remaining_cost
                    print("未实现盈亏: {0:.2f}".format(unrealized_profit))
                    
                    # 计算总盈亏（已实现 + 未实现）
                    total_profit_with_unrealized = total_profit + unrealized_profit
                    print("总盈亏(含未实现): {0:.2f}".format(total_profit_with_unrealized))
                else:
                    print("警告: 无法获取最后价格，剩余股票将不计入XIRR计算")
            
            # 计算总现金流
            total_inflow = total_sell_value
            total_outflow = total_buy_value
            
            if total_remaining_shares > 0 and last_price:
                total_inflow += remaining_value
            
            total_cash_flow = total_inflow - total_outflow
            print("\n总现金流: 流入 {0:.2f} - 流出 {1:.2f} = {2:.2f}".format(
                total_inflow, total_outflow, total_cash_flow))
            
            # 按日期排序现金流
            cash_flows.sort(key=lambda x: x[0])
            
            # 打印现金流用于调试
            print("\n排序后的现金流:")
            for date, amount in cash_flows:
                print(f"{date}\t{amount:.2f}")
            
            # 计算XIRR
            print("\n计算XIRR...")
            if len(cash_flows) < 2:
                print("错误: 现金流数量不足，无法计算XIRR")
                self.db_connector.release_connection(conn)
                return None
            
            # 分离日期和金额列表
            dates = [cf[0] for cf in cash_flows]
            amounts = [cf[1] for cf in cash_flows]
            
            # 计算XIRR
            xirr = self.calculate_xirr(dates, amounts, guess=0.06)  # 使用0.06作为初始猜测值
            if xirr is not None:
                xirr_value = xirr * 100  # 转换为百分比
                print("XIRR计算结果: {0:.2f}%".format(xirr_value))
            else:
                print("XIRR计算失败")
                xirr = 0  # 设置默认值，避免后续操作出错
                xirr_value = 0
            
            # 保存XIRR结果到数据库
            print("\n保存XIRR结果到数据库...")
            
            # 检查是否已存在记录
            cursor.execute("""
                SELECT id FROM backtest_xirr 
                WHERE backtest_id = %s
            """, (backtest_id,))
            
            existing_record = cursor.fetchone()
            
            if existing_record:
                # 更新现有记录
                cursor.execute("""
                    UPDATE backtest_xirr 
                    SET xirr = %s, 
                        xirr_value = %s,
                        total_buy_value = %s, 
                        total_sell_value = %s, 
                        remaining_shares = %s, 
                        remaining_value = %s, 
                        total_cash_flow = %s, 
                        calculation_time = NOW(),
                        has_incomplete_trades = %s,
                        notes = %s
                    WHERE backtest_id = %s
                """, (
                    float(xirr) if xirr is not None else None,
                    float(xirr_value) if xirr is not None else None,
                    float(total_buy_value),
                    float(total_sell_value),
                    int(total_remaining_shares),
                    float(remaining_value) if total_remaining_shares > 0 and last_price else 0.0,
                    float(total_cash_flow),
                    bool(has_incomplete_trades),
                    f"交易专用XIRR计算，剩余股数{total_remaining_shares}，Excel兼容计算方式",
                    int(backtest_id)
                ))
            else:
                # 插入新记录
                cursor.execute("""
                    INSERT INTO backtest_xirr 
                        (backtest_id, xirr, xirr_value, xirr_type, 
                         total_buy_value, total_sell_value, 
                         remaining_shares, remaining_value, total_cash_flow, 
                         calculation_time, has_incomplete_trades, notes)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), %s, %s)
                """, (
                    int(backtest_id),
                    float(xirr) if xirr is not None else None,
                    float(xirr_value) if xirr is not None else None,
                    'trades_only',
                    float(total_buy_value),
                    float(total_sell_value),
                    int(total_remaining_shares),
                    float(remaining_value) if total_remaining_shares > 0 and last_price else 0.0,
                    float(total_cash_flow),
                    bool(has_incomplete_trades),
                    f"交易专用XIRR计算，剩余股数{total_remaining_shares}，Excel兼容计算方式"
                ))
            
            conn.commit()
            
            # 关闭数据库连接
            cursor.close()
            self.db_connector.release_connection(conn)
            
            return {
                'xirr': xirr,
                'xirr_value': xirr_value,  # 直接返回百分比值
                'total_buy_value': total_buy_value,
                'total_sell_value': total_sell_value,
                'remaining_shares': total_remaining_shares,
                'remaining_value': remaining_value if total_remaining_shares > 0 and last_price else 0,
                'total_cash_flow': total_cash_flow
            }
            
        except Exception as e:
            print("计算XIRR时出错: {0}".format(str(e)))
            import traceback
            traceback.print_exc()
            if conn:
                self.db_connector.release_connection(conn)
            return None
            
    def export_to_excel(self, backtest_id, file_path=None):
        """导出XIRR计算结果到Excel
        
        Args:
            backtest_id: 回测ID
            file_path: 导出文件路径，如果为None则使用默认路径
            
        Returns:
            导出文件路径
        """
        try:
            # 获取XIRR计算结果
            xirr_dict = self.calculate_backtest_xirr(backtest_id)
            if not xirr_dict:
                print("无法获取XIRR计算结果")
                return None
            
            # 获取数据库连接
            conn = self.db_connector.get_connection()
            cursor = conn.cursor()
            
            # 获取回测基本信息
            cursor.execute("""
                SELECT id, stock_code, start_date, end_date, initial_capital, final_capital, total_profit, total_profit_rate
                FROM backtest_results
                WHERE id = %s
            """, (backtest_id,))
            
            backtest_info = cursor.fetchone()
            if not backtest_info:
                print("未找到回测ID: {0}的信息".format(backtest_id))
                self.db_connector.release_connection(conn)
                return None
            
            # 转换为字典
            backtest_info_dict = {
                'id': backtest_info[0],
                'stock_code': backtest_info[1],
                'start_date': backtest_info[2],
                'end_date': backtest_info[3],
                'initial_capital': float(backtest_info[4]),
                'final_capital': float(backtest_info[5]),
                'total_profit': float(backtest_info[6]),
                'total_profit_rate': float(backtest_info[7])
            }
            
            # 获取配对交易记录
            cursor.execute("""
                SELECT id, backtest_id, level, grid_type, 
                       buy_time, buy_price, buy_amount, buy_value, 
                       sell_time, sell_price, sell_amount, sell_value,
                       remaining, band_profit, band_profit_rate, status
                FROM backtest_paired_trades 
                WHERE backtest_id = %s
                ORDER BY buy_time
            """, (backtest_id,))
            
            paired_trades = cursor.fetchall()
            
            # 准备现金流数据
            cash_flows = []  # 存储(日期, 金额, 备注)对
            buy_cash_flows = []  # 买入现金流
            sell_cash_flows = []  # 卖出现金流
            
            # 计算剩余持仓的价值
            partially_sold_remaining_shares = 0  # 部分卖出后剩余的股数
            unsold_shares = 0  # 完全未卖出的股数
            total_remaining_shares = 0  # 总剩余股数
            
            # 添加买入交易的现金流（负值）
            for trade in paired_trades:
                buy_time = trade[4]
                buy_price = float(trade[5])
                buy_amount = int(trade[6])
                buy_value = float(trade[7])
                sell_time = trade[8]
                sell_price = float(trade[9]) if trade[9] is not None else 0
                sell_amount = int(trade[10]) if trade[10] is not None else 0
                sell_value = float(trade[11]) if trade[11] is not None else 0
                remaining = int(trade[12]) if trade[12] is not None else 0
                
                # 买入是负现金流
                buy_cash_flows.append((buy_time, buy_value))
                cash_flows.append((buy_time, -buy_value, "买入"))
                
                # 只添加有卖出的交易
                if sell_time and sell_value > 0:
                    sell_cash_flows.append((sell_time, sell_value))
                    cash_flows.append((sell_time, sell_value, "卖出"))
                
                # 计算两种类型的剩余股数
                if sell_amount > 0 and remaining > 0:
                    # 部分卖出的交易
                    partially_sold_remaining_shares += remaining
                elif sell_amount == 0:
                    # 完全未卖出的交易
                    unsold_shares += buy_amount
                
                # 累计总剩余股数
                total_remaining_shares += (buy_amount - sell_amount)
            
            # 获取最后价格
            last_price = None
            cursor.execute("""
                SELECT sell_price FROM backtest_paired_trades 
                WHERE backtest_id = %s AND sell_price IS NOT NULL 
                ORDER BY sell_time DESC LIMIT 1
            """, (backtest_id,))
            result = cursor.fetchone()
            if result:
                last_price = float(result[0])
            
            # 添加剩余持仓的现金流
            if total_remaining_shares > 0 and last_price:
                end_date = backtest_info_dict['end_date']
                
                # 分别添加两种类型的剩余股票
                if partially_sold_remaining_shares > 0:
                    partially_sold_value = partially_sold_remaining_shares * last_price
                    cash_flows.append((end_date, partially_sold_value, f"买入卖出后剩余底仓: {partially_sold_remaining_shares} 份额 (价格: {last_price:.4f})"))
                
                if unsold_shares > 0:
                    unsold_value = unsold_shares * last_price
                    cash_flows.append((end_date, unsold_value, f"买入未卖出底仓: {unsold_shares} 份额 (价格: {last_price:.4f})"))
            
            # 创建Excel工作簿
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws_summary = wb.active
            ws_summary.title = "XIRR汇总"
            
            # 添加标题
            ws_summary.append(["交易专用XIRR计算结果 - 回测ID: {0}".format(backtest_id)])
            ws_summary.append(["基金代码", backtest_info_dict['stock_code']])
            ws_summary.append(["开始日期", backtest_info_dict['start_date']])
            ws_summary.append(["结束日期", backtest_info_dict['end_date']])
            
            # 添加XIRR结果
            ws_summary.append(["XIRR(年化收益率)", f"{xirr_dict['xirr_value']:.2f}%"])
            ws_summary.append(["总买入金额", f"{xirr_dict['total_buy_value']:.2f}"])
            ws_summary.append(["总卖出金额", f"{xirr_dict['total_sell_value']:.2f}"])
            ws_summary.append(["剩余股数", f"{total_remaining_shares}"])
            ws_summary.append(["买入卖出后剩余底仓", f"{partially_sold_remaining_shares}"])
            ws_summary.append(["买入未卖出底仓", f"{unsold_shares}"])
            
            # 计算剩余股票估值
            remaining_value = total_remaining_shares * last_price if total_remaining_shares > 0 and last_price > 0 else 0
            ws_summary.append(["剩余股票估值", f"{remaining_value:.2f}"])
            
            # 计算总现金流
            total_cash_flow = xirr_dict['total_sell_value'] + remaining_value - xirr_dict['total_buy_value']
            ws_summary.append(["总现金流", f"{total_cash_flow:.2f}"])
            
            # 检查是否有未完成交易
            has_incomplete_trades = total_remaining_shares > 0
            if has_incomplete_trades:
                ws_summary.append(["注意", "存在未完成交易，XIRR计算结果包含当前持仓价值"])
                
            # 创建现金流工作表
            ws_cashflow = wb.create_sheet(title="现金流数据")
            
            # 添加现金流表头
            ws_cashflow.append(["date", "amount", "备注"])
            
            # 按日期排序
            all_cash_flows = sorted(cash_flows, key=lambda x: x[0])
            
            # 添加到工作表
            for date, amount, remark in all_cash_flows:
                ws_cashflow.append([date, amount, remark])
                
            # 添加XIRR公式
            row_count = len(all_cash_flows) + 2  # 表头 + 所有现金流 + 1
            ws_cashflow.append([])
            ws_cashflow.append(["Excel XIRR公式", f"=XIRR(B2:B{row_count-1},A2:A{row_count-1})", "使用Excel的XIRR公式计算年化收益率"])
            ws_cashflow.append(["Excel XIRR结果", "", "请点击此单元格，Excel会自动计算XIRR结果"])
            
            # 设置单元格格式
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 设置表头格式
            for cell in ws_cashflow[1]:
                cell.font = Font(bold=True)
                
            # 设置XIRR公式行格式
            for cell in ws_cashflow[row_count+1]:
                cell.font = Font(bold=True)
            
            # 设置XIRR结果行格式
            for cell in ws_cashflow[row_count+2]:
                cell.font = Font(bold=True)
                
            # 创建持仓明细工作表
            ws_holdings = wb.create_sheet(title="持仓明细")
            
            # 添加持仓明细表头
            ws_holdings.append(["交易ID", "买入时间", "买入价格", "买入数量", "买入金额", 
                               "卖出时间", "卖出价格", "卖出数量", "卖出金额", "剩余数量", "状态"])
            
            # 添加持仓明细数据
            for trade in paired_trades:
                trade_id = trade[0]
                buy_time = trade[4]
                buy_price = float(trade[5])
                buy_amount = int(trade[6])
                buy_value = float(trade[7])
                sell_time = trade[8] if trade[8] else ""
                sell_price = float(trade[9]) if trade[9] is not None else 0
                sell_amount = int(trade[10]) if trade[10] is not None else 0
                sell_value = float(trade[11]) if trade[11] is not None else 0
                remaining = int(trade[12]) if trade[12] is not None else 0
                status = trade[15]
                
                ws_holdings.append([trade_id, buy_time, buy_price, buy_amount, buy_value,
                                  sell_time, sell_price, sell_amount, sell_value, remaining, status])
            
            # 设置表头格式
            for cell in ws_holdings[1]:
                cell.font = Font(bold=True)
            
            # 保存Excel文件
            if file_path is None:
                import os
                file_path = os.path.join(os.getcwd(), f"交易专用XIRR_{backtest_id}.xlsx")
            
            wb.save(file_path)
            print("Excel文件已保存: {0}".format(file_path))
            
            # 关闭数据库连接
            cursor.close()
            self.db_connector.release_connection(conn)
            
            return file_path
        except Exception as e:
            print("导出Excel时出错: {0}".format(str(e)))
            traceback.print_exc()
            if 'conn' in locals() and conn:
                self.db_connector.release_connection(conn)
            return None 